import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import os, hashlib

# --------------------------------------- CONFIG ---------------------------------------
BG_COLOR = '#1A1A2E'
FG_COLOR = '#F0A500'
CARD_BG = '#243447'
BTN_BG = '#F0A500'
BTN_FG = '#1A1A2E'
ENTRY_BG = '#2E2E3C'
ENTRY_FG = 'white'
BTN_WIDTH = 20
BTN_HEIGHT = 2

EXCEL_FILE = 'animal_data.xlsx'
ADMIN_USER = 'admin'
ADMIN_PASS_HASH = hashlib.sha256('1234'.encode()).hexdigest()

# ----------------------------------- UTIL FUNCTIONS ----------------------------------
def ensure_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Animals'
        ws.append(['ID','Name','Species','Age','Description','Photo'])
        wb.create_sheet('Users')
        wb['Users'].append(['Username','PasswordHash','Email'])
        wb.create_sheet('Adoptions')
        wb['Adoptions'].append(['Animal ID','Animal Name','Adopter Username','Adopter Email'])
        wb.create_sheet('AdoptionRequests')
        wb['AdoptionRequests'].append(['Animal ID','Animal Name','Username','User Email'])
        wb.save(EXCEL_FILE)


def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def load_image(path, size=(200, 200)):
    if path and os.path.exists(path):
        try:
            img = Image.open(path).resize(size)
            return ImageTk.PhotoImage(img)
        except Exception as e:
            print(f"Failed to load image {path}: {e}")
            return None
    else:
        print(f"Image path not found: {path}")
        return None



def confirm(title, msg):
    return messagebox.askyesno(title, msg)

# ------------------------------------ MAIN APP ------------------------------------
class AnimalRescueHub:
    def __init__(self, root):
        self.root = root
        root.title('Animal Rescue & Adoption Hub')
        root.state('zoomed')
        root.configure(bg=BG_COLOR)
        ensure_excel()
        self.current_user = None
        self.setup_style()
        self.show_main_menu()
        self.img_refs = []


    def setup_style(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TButton', background=BTN_BG, foreground=BTN_FG, font=('Arial', 12, 'bold'))
        style.map('TButton', background=[('active', FG_COLOR)])
        style.configure('TEntry', fieldbackground=ENTRY_BG, foreground=ENTRY_FG)
        style.configure('Treeview', background=CARD_BG, fieldbackground=CARD_BG, foreground='white', rowheight=25)
        style.configure('Treeview.Heading', font=('Arial', 14, 'bold'), background=CARD_BG, foreground=FG_COLOR)

    def clear(self):
        for w in self.root.winfo_children(): w.destroy()

    # --------------------------- Main Menu ---------------------------
    def show_main_menu(self):
        self.clear()
        lbl = tk.Label(self.root, text='🐾 Animal Adoption Hub 🐾', bg=BG_COLOR, fg=FG_COLOR, font=('Arial', 36, 'bold'))
        lbl.pack(pady=60)
        for txt, cmd in [('Admin Login', self.show_admin_login), ('User Login', self.show_user_login), ('User Register', self.show_user_register)]:
            btn = ttk.Button(self.root, text=txt, command=cmd)
            btn.pack(pady=20, ipadx=30, ipady=10)

    # --------------------------- Admin Flow ---------------------------
    def show_admin_login(self):
        self.clear()
        tk.Label(self.root, text='Admin Login', bg=BG_COLOR, fg=FG_COLOR, font=('Arial',30,'bold')).pack(pady=40)
        self.admin_user = self.create_entry('Username')
        self.admin_pass = self.create_entry('Password', show='*')
        ttk.Button(self.root, text='Login', command=self.check_admin).pack(pady=20)
        ttk.Button(self.root, text='⬅️ Back', command=self.show_main_menu).pack(pady=10)

    def check_admin(self):
        if self.admin_user.get()==ADMIN_USER and hash_pw(self.admin_pass.get())==ADMIN_PASS_HASH:
            self.show_admin_dashboard()
        else:
            messagebox.showerror('Error','Invalid credentials')

    def show_admin_dashboard(self):
        self.clear()
        # notify pending requests
        wb = load_workbook(EXCEL_FILE); cnt = wb['AdoptionRequests'].max_row-1
        if cnt>0: messagebox.showinfo('Pending Requests',f'You have {cnt} adoption request(s)')
        tk.Label(self.root, text='Admin Dashboard', bg=BG_COLOR, fg=FG_COLOR, font=('Arial',32,'bold')).pack(pady=30)
        for txt, cmd in [('Add Animal', self.add_animal), 
                         ('View Animals', self.view_animals), 
                         ('Delete Animal', self.delete_animal), 
                         ('Upload Photo', self.upload_photo), 
                         ('View Requests', self.view_adoption_requests), 
                         ('View Adoptions', self.view_adoptions),
                           ('Logout', self.logout)]:
            ttk.Button(self.root, text=txt, command=cmd, width=BTN_WIDTH).pack(pady=10, ipadx=20)

    def logout(self):
        if confirm('Logout','Confirm logout?'): self.show_main_menu()

    # --------------------------- Animal Management ---------------------------
    def add_animal(self):
        w=tk.Toplevel(self.root); w.title('Add Animal'); w.configure(bg=BG_COLOR)
        entries={}
        for field in ['Name','Species','Age','Description']:
            tk.Label(w,text=field, bg=BG_COLOR, fg=FG_COLOR).pack(pady=5)
            entries[field] = ttk.Entry(w)
            entries[field].pack(pady=5)
        def save():
            wb=load_workbook(EXCEL_FILE); ws=wb['Animals']; nid=ws.max_row
            vals=[entries[f].get() for f in entries]
            if all(vals): ws.append([nid]+vals+['']); wb.save(EXCEL_FILE); messagebox.showinfo('Added','Animal added'); w.destroy()
            else: messagebox.showerror('Error','Fill all fields')
        ttk.Button(w,text='Save', command=save).pack(pady=20)
        ttk.Button(w,text='⬅️ Back', command=w.destroy).pack(pady=5)

    def view_animals(self):
        w = tk.Toplevel(self.root)
        w.title('All Animals')
        w.state('zoomed')
        w.configure(bg=BG_COLOR)

        tk.Label(w, text='All Animals', bg=BG_COLOR, fg=FG_COLOR, font=('Arial', 40, 'bold')).pack(pady=30)

        frame = tk.Frame(w, bg=BG_COLOR)
        frame.pack(fill='both', expand=True)

        canvas = tk.Canvas(frame, bg=BG_COLOR)
        sb = ttk.Scrollbar(frame, orient='vertical', command=canvas.yview)
        cont = tk.Frame(canvas, bg=BG_COLOR)
        cont.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=cont, anchor='nw')
        canvas.configure(yscrollcommand=sb.set)

        canvas.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        self.img_refs.clear()  # clear any old images before loading new ones

        wb = load_workbook(EXCEL_FILE)
        ws = wb['Animals']

        r, c = 0, 0
        for aid, name, species, age, desc, photo in ws.iter_rows(min_row=2, values_only=True):
            card = tk.Frame(cont, bg=CARD_BG, bd=3, relief='ridge')
            card.grid(row=r, column=c, padx=20, pady=20)

            # Animal image
            if photo and os.path.exists(photo):
                img = load_image(photo, (200, 200))
                if img:
                    img_label = tk.Label(card, image=img, bg=CARD_BG)
                    img_label.pack(pady=10)
                    self.img_refs.append(img)  # keep reference
            else:
                tk.Label(card, text='No Photo Available', bg=CARD_BG, fg=FG_COLOR, font=('Arial', 14, 'italic')).pack(pady=10)

            # Animal details
            tk.Label(card, text=f'ID: {aid}', bg=CARD_BG, fg=FG_COLOR, font=('Arial', 16, 'bold')).pack(pady=3)
            tk.Label(card, text=f'Name: {name}', bg=CARD_BG, fg=FG_COLOR, font=('Arial', 14)).pack(pady=2)
            tk.Label(card, text=f'Species: {species}', bg=CARD_BG, fg=FG_COLOR, font=('Arial', 14)).pack(pady=2)
            tk.Label(card, text=f'Age: {age}', bg=CARD_BG, fg=FG_COLOR, font=('Arial', 14)).pack(pady=2)
            tk.Label(card, text=f'Description: {desc}', bg=CARD_BG, fg=FG_COLOR, wraplength=250, justify='center').pack(pady=5)

            c += 1
            if c >= 6:  # 2 cards per row
                c = 0
                r += 1

        ttk.Button(w, text='⬅️ Back', command=w.destroy).pack(pady=20)


    def delete_animal(self):
        if not confirm('Delete','Delete selected animal?'): return
        aid=simpledialog.askinteger('Delete','Animal ID:')
        if not aid: return
        wb=load_workbook(EXCEL_FILE); ws=wb['Animals']
        for row in ws.iter_rows(min_row=2):
            if row[0].value==aid: ws.delete_rows(row[0].row); wb.save(EXCEL_FILE); messagebox.showinfo('Deleted','Animal removed'); return
        messagebox.showerror('Error','ID not found')

    def upload_photo(self):
        aid=simpledialog.askinteger('Photo','Animal ID:'); wb=load_workbook(EXCEL_FILE); ws=wb['Animals']
        for row in ws.iter_rows(min_row=2):
            if row[0].value==aid:
                path=filedialog.askopenfilename(filetypes=[('Image','*.png *.jpg')])
                if path: row[5].value=path; wb.save(EXCEL_FILE); messagebox.showinfo('Done','Photo added')
                return
        messagebox.showerror('Error','ID not found')

    # --------------------------- Adoption Requests ---------------------------
    def view_adoption_requests(self):
        w = tk.Toplevel(self.root)
        w.title('Adoption Requests')
        w.state('zoomed')
        w.configure(bg=BG_COLOR)

        tk.Label(w, text='Adoption Requests', bg=BG_COLOR, fg=FG_COLOR, font=('Arial', 40, 'bold')).pack(pady=30)

        frame = tk.Frame(w, bg=BG_COLOR)
        frame.pack(fill='both', expand=True)

        canvas = tk.Canvas(frame, bg=BG_COLOR)
        sb = ttk.Scrollbar(frame, orient='vertical', command=canvas.yview)
        cont = tk.Frame(canvas, bg=BG_COLOR)
        cont.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=cont, anchor='nw')
        canvas.configure(yscrollcommand=sb.set)

        canvas.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        self.img_refs.clear()
        wb = load_workbook(EXCEL_FILE)
        wsA = wb['Animals']
        wsR = wb['AdoptionRequests']

        animal_photos = {r[0]: r[5] for r in wsA.iter_rows(min_row=2, values_only=True)}

        r, c = 0, 0
        def process_request(aid, name, user, email, action):
            wsAd = wb['Adoptions']
            if action == 'accept':
                wsAd.append([aid, name, user, email])
                messagebox.showinfo('Accepted', f'{name} adopted by {user}')
            for row in wsR.iter_rows(min_row=2):
                if row[0].value == aid and row[2].value == user:
                    wsR.delete_rows(row[0].row)
                    break
            wb.save(EXCEL_FILE)
            w.destroy()
            self.view_adoption_requests()

        for aid, name, user, email in wsR.iter_rows(min_row=2, values_only=True):
            card = tk.Frame(cont, bg=CARD_BG, bd=3, relief='ridge')
            card.grid(row=r, column=c, padx=10, pady=10)

            photo = animal_photos.get(aid)
            img = load_image(photo, (180, 180)) if photo and os.path.exists(photo) else None
            if img:
                img_label = tk.Label(card, image=img, bg=CARD_BG)
                img_label.pack(pady=5)
                self.img_refs.append(img)
            else:
                tk.Label(card, text='No Photo', bg=CARD_BG, fg=FG_COLOR).pack(pady=5)

            tk.Label(card, text=f'ID: {aid}', bg=CARD_BG, fg=FG_COLOR).pack()
            tk.Label(card, text=f'Name: {name}', bg=CARD_BG, fg=FG_COLOR).pack()
            tk.Label(card, text=f'User: {user}', bg=CARD_BG, fg=FG_COLOR).pack()
            tk.Label(card, text=f'Email: {email}', bg=CARD_BG, fg=FG_COLOR).pack()

            ttk.Button(card, text='Accept', command=lambda a=aid, n=name, u=user, e=email: process_request(a, n, u, e, 'accept')).pack(pady=3)
            ttk.Button(card, text='Reject', command=lambda a=aid, n=name, u=user, e=email: process_request(a, n, u, e, 'reject')).pack(pady=3)

            c += 1
            if c >= 6:
                c = 0
                r += 1

        ttk.Button(w, text='⬅️ Back', command=w.destroy).pack(pady=20)


    def view_adoptions(self):
        w = tk.Toplevel(self.root)
        w.title('Adoptions')
        w.state('zoomed')
        w.configure(bg=BG_COLOR)

        tk.Label(w, text='Completed Adoptions', bg=BG_COLOR, fg=FG_COLOR, font=('Arial', 40, 'bold')).pack(pady=30)

        frame = tk.Frame(w, bg=BG_COLOR)
        frame.pack(fill='both', expand=True)

        canvas = tk.Canvas(frame, bg=BG_COLOR)
        sb = ttk.Scrollbar(frame, orient='vertical', command=canvas.yview)
        cont = tk.Frame(canvas, bg=BG_COLOR)
        cont.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=cont, anchor='nw')
        canvas.configure(yscrollcommand=sb.set)

        canvas.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        self.img_refs.clear()
        wb = load_workbook(EXCEL_FILE)
        wsA = wb['Animals']
        wsAd = wb['Adoptions']

        animal_photos = {r[0]: r[5] for r in wsA.iter_rows(min_row=2, values_only=True)}

        r, c = 0, 0
        for aid, name, user, email in wsAd.iter_rows(min_row=2, values_only=True):
            card = tk.Frame(cont, bg=CARD_BG, bd=3, relief='ridge')
            card.grid(row=r, column=c, padx=10, pady=10)

            photo = animal_photos.get(aid)
            img = load_image(photo, (180, 180)) if photo and os.path.exists(photo) else None
            if img:
                img_label = tk.Label(card, image=img, bg=CARD_BG)
                img_label.pack(pady=5)
                self.img_refs.append(img)
            else:
                tk.Label(card, text='No Photo', bg=CARD_BG, fg=FG_COLOR).pack(pady=5)

            tk.Label(card, text=f'ID: {aid}', bg=CARD_BG, fg=FG_COLOR).pack()
            tk.Label(card, text=f'Name: {name}', bg=CARD_BG, fg=FG_COLOR).pack()
            tk.Label(card, text=f'User: {user}', bg=CARD_BG, fg=FG_COLOR).pack()
            tk.Label(card, text=f'Email: {email}', bg=CARD_BG, fg=FG_COLOR).pack()

            c += 1
            if c >= 6:
                c = 0
                r += 1

        ttk.Button(w, text='⬅️ Back', command=w.destroy).pack(pady=20)


    # --------------------------- User Flow ---------------------------
    def show_user_register(self):
        self.clear(); tk.Label(self.root,text='Register',bg=BG_COLOR,fg=FG_COLOR,font=('Arial',30,'bold')).pack(pady=30)
        self.r_u=self.create_entry('Username'); self.r_e=self.create_entry('Email'); self.r_p=self.create_entry('Password',show='*')
        ttk.Button(self.root,text='Register',command=self.register_user).pack(pady=20)
        ttk.Button(self.root,text='⬅️ Back',command=self.show_main_menu).pack(pady=10)

    def register_user(self):
        u,e,p=self.r_u.get(),self.r_e.get(),self.r_p.get()
        if all([u,e,p]):
            wb=load_workbook(EXCEL_FILE); ws=wb['Users']; ws.append([u,hash_pw(p),e]); wb.save(EXCEL_FILE)
            messagebox.showinfo('Done','Registered'); self.show_main_menu()
        else: messagebox.showerror('Error','Fill all')

    def show_user_login(self):
        self.clear(); tk.Label(self.root,text='User Login',bg=BG_COLOR,fg=FG_COLOR,font=('Arial',30,'bold')).pack(pady=30)
        self.l_u=self.create_entry('Username'); self.l_p=self.create_entry('Password',show='*')
        ttk.Button(self.root,text='Login',command=self.check_user).pack(pady=20)
        ttk.Button(self.root,text='⬅️ Back',command=self.show_main_menu).pack(pady=10)

    def check_user(self):
        u,p=self.l_u.get(),hash_pw(self.l_p.get()); ws=load_workbook(EXCEL_FILE)['Users']
        for r in ws.iter_rows(min_row=2,values_only=True):
            if r[0]==u and r[1]==p: self.current_user=(u,r[2]); self.show_user_dashboard(); return
        messagebox.showerror('Error','Invalid cred')

    def show_user_dashboard(self):
        self.clear(); tk.Label(self.root,text=f'Welcome {self.current_user[0]}',bg=BG_COLOR,fg=FG_COLOR,font=('Arial',32,'bold')).pack(pady=30)
        ws=load_workbook(EXCEL_FILE)['Adoptions']; adopted=[(r[0],r[1]) for r in ws.iter_rows(min_row=2,values_only=True) if r[2]==self.current_user[0]]
        if adopted:
            tk.Label(self.root,text='Your Adoptions',bg=BG_COLOR,fg=FG_COLOR,font=('Arial',24,'bold')).pack(pady=10)
            for aid,name in adopted: tk.Label(self.root,text=f'ID:{aid} {name}',bg=BG_COLOR,fg=FG_COLOR,font=('Arial',20)).pack(pady=5)
        else:
            tk.Label(self.root,text='No adoptions yet',bg=BG_COLOR,fg=FG_COLOR,font=('Arial',20)).pack(pady=10)
        ttk.Button(self.root,text='Request Adoption',command=self.adopt_dialog).pack(pady=20)
        ttk.Button(self.root,text='⬅️ Logout',command=self.show_main_menu).pack(pady=10)

    def adopt_dialog(self):
    # 1) Clear out any existing widgets on the main window
        self.clear()
        self.root.title('Choose Animal')
        self.root.state('zoomed')
        self.root.configure(bg=BG_COLOR)

    # 2) Header
        tk.Label(self.root, text='Choose Animal', bg=BG_COLOR, fg=FG_COLOR,
             font=('Arial', 28, 'bold')).pack(pady=20)

    # 3) Scrollable card container
        frame = tk.Frame(self.root, bg=BG_COLOR)
        frame.pack(fill='both', expand=True)
        canvas = tk.Canvas(frame, bg=BG_COLOR)
        sb = ttk.Scrollbar(frame, orient='vertical', command=canvas.yview)
        cont = tk.Frame(canvas, bg=BG_COLOR)
        cont.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=cont, anchor='nw')
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

    # 4) Load data
        wb = load_workbook(EXCEL_FILE)
        wsA = wb['Adoptions']
        wsR = wb['AdoptionRequests']
        ws  = wb['Animals']
        adopted   = {r[0] for r in wsA.iter_rows(min_row=2, values_only=True)}
        requested = {(r[0], r[2]) for r in wsR.iter_rows(min_row=2, values_only=True)}

    # 5) Build cards
        r = c = 0
        self.img_refs.clear()
        for aid, name, species, age, desc, photo in ws.iter_rows(min_row=2, values_only=True):
            card = tk.Frame(cont, bg=CARD_BG, bd=3, relief='ridge')
            card.grid(row=r, column=c, padx=10, pady=10)

            # Photo
            img = load_image(photo, (200,200))
            if img:
                tk.Label(card, image=img, bg=CARD_BG).pack(pady=5)
                self.img_refs.append(img)
            else:
                tk.Label(card, text='No Photo', bg=CARD_BG, fg=FG_COLOR).pack(pady=20)

            # Details
            tk.Label(card, text=f'{name} ({species})', bg=CARD_BG, fg=FG_COLOR).pack()
            tk.Label(card, text=f'Age: {age}', bg=CARD_BG, fg=FG_COLOR).pack()
            tk.Label(card, text=f'Description: {desc}', bg=CARD_BG, fg=FG_COLOR,
                 wraplength=200, justify='center').pack(pady=5)

            # Request / Status
            if aid in adopted:
                status = 'Adopted'
            elif (aid, self.current_user[0]) in requested:
                status = 'Pending'
            else:
                status = ''

            if not status:
            # Pass only aid & name to send_request
                btn = ttk.Button(card,
                             text='Request',
                             command=lambda a=aid, n=name: self.send_request(a, n))
                btn.pack(pady=5)
            else:
                tk.Label(card, text=status, bg=CARD_BG, fg='red').pack(pady=5)

            c += 1
            if c >= 6:
                c = 0
                r += 1

    # 6) Back button
        ttk.Button(self.root, text='⬅️ Back', command=self.show_user_dashboard).pack(pady=20)




    def send_request(self, aid, name):
        wb = load_workbook(EXCEL_FILE)
        ws = wb['AdoptionRequests']
        ws.append([aid, name, self.current_user[0], self.current_user[1]])
        wb.save(EXCEL_FILE)
        messagebox.showinfo('Requested', f'Request sent for {name}')
        # Navigate back to the user dashboard
        self.show_user_dashboard()



    # helper to create entries
    def create_entry(self, placeholder, show=None):
        ent=ttk.Entry(self.root, show=show if show else '')
        ent.insert(0, placeholder)
        ent.pack(pady=5)
        def on_in(e):
            if ent.get()==placeholder: ent.delete(0,tk.END)
        def on_out(e):
            if not ent.get(): ent.insert(0, placeholder)
        ent.bind('<FocusIn>',on_in); ent.bind('<FocusOut>',on_out)
        return ent

if __name__=='__main__':
    root=tk.Tk(); 
    app=AnimalRescueHub(root); 
    root.mainloop()
